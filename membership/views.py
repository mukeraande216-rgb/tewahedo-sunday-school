from datetime import date, datetime

from django.core.paginator import Paginator
from django.contrib import messages
from django.contrib.admin.views.decorators import staff_member_required
from django.db import transaction
from django.db.models import Q
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.utils.dateparse import parse_date
from django.views.decorators.http import require_POST
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.datetime import from_excel

from .forms import (
    MemberForm,
    MemberImportForm,
    MembershipApplicationForm,
)
from .models import Member, MembershipApplication

HEADER_ALIASES = {
    "member id": "member_id",
    "memberid": "member_id",
    "full name": "full_name",
    "name": "full_name",
    "baptism name": "baptism_name",
    "address": "address",
    "yenesha abat": "yenesha_abat",
    "yeneshe abat": "yenesha_abat",
    "date": "membership_date",
    "membership date": "membership_date",
    "phone number": "phone_number",
    "phone": "phone_number",
    "active member": "is_active",
    "active": "is_active",
    "notes": "notes",
}


def _normalize_header(value):
    text = str(value or "").strip().lower()
    text = text.replace("_", " ").replace("-", " ")
    return " ".join(text.split())


def _clean_text(value):
    if value is None:
        return ""

    if isinstance(value, float) and value.is_integer():
        return str(int(value))

    return str(value).strip()


def _parse_membership_date(value, workbook_epoch):
    if value in (None, ""):
        return None

    if isinstance(value, datetime):
        return value.date()

    if isinstance(value, date):
        return value

    if isinstance(value, (int, float)):
        try:
            converted = from_excel(value, epoch=workbook_epoch)

            if isinstance(converted, datetime):
                return converted.date()

            if isinstance(converted, date):
                return converted
        except (TypeError, ValueError, OverflowError):
            return None

    text = _clean_text(value)

    parsed = parse_date(text)
    if parsed:
        return parsed

    accepted_formats = (
        "%m/%d/%Y",
        "%m/%d/%y",
        "%Y/%m/%d",
        "%m-%d-%Y",
        "%m-%d-%y",
    )

    for date_format in accepted_formats:
        try:
            return datetime.strptime(text, date_format).date()
        except ValueError:
            continue

    return None


def _parse_active_status(value):
    if value in (None, ""):
        return True

    if isinstance(value, bool):
        return value

    text = _clean_text(value).lower()

    return text not in {
        "no",
        "n",
        "false",
        "0",
        "inactive",
    }


@staff_member_required
def member_list(request):
    search_query = request.GET.get("q", "").strip()

    all_members = Member.objects.all()

    total_count = all_members.count()
    active_count = all_members.filter(is_active=True).count()
    inactive_count = all_members.filter(is_active=False).count()

    applications = MembershipApplication.objects.all()

    pending_application_count = applications.filter(
        status=MembershipApplication.Status.PENDING
    ).count()

    under_review_application_count = applications.filter(
        status=MembershipApplication.Status.UNDER_REVIEW
    ).count()

    applications_needing_review_count = (
        pending_application_count
        + under_review_application_count
    )

    members = all_members

    if search_query:
        members = members.filter(
            Q(member_id__icontains=search_query)
            | Q(full_name__icontains=search_query)
            | Q(baptism_name__icontains=search_query)
            | Q(phone_number__icontains=search_query)
            | Q(address__icontains=search_query)
            | Q(yenesha_abat__icontains=search_query)
        )

    context = {
        "members": members,
        "search_query": search_query,
        "total_count": total_count,
        "active_count": active_count,
        "inactive_count": inactive_count,
        "pending_application_count": pending_application_count,
        "under_review_application_count": (
            under_review_application_count
        ),
        "applications_needing_review_count": (
            applications_needing_review_count
        ),
    }

    return render(
        request,
        "membership/member_list.html",
        context,
    )

@staff_member_required
def membership_application_list(request):
    search_query = request.GET.get("q", "").strip()
    status_filter = request.GET.get("status", "").strip()

    applications = MembershipApplication.objects.select_related(
        "reviewed_by",
        "approved_member",
    )

    if search_query:
        applications = applications.filter(
            Q(application_id__icontains=search_query)
            | Q(full_name__icontains=search_query)
            | Q(baptism_name__icontains=search_query)
            | Q(phone_number__icontains=search_query)
            | Q(email__icontains=search_query)
            | Q(address__icontains=search_query)
        )

    valid_statuses = {
        value
        for value, label in MembershipApplication.Status.choices
    }

    if status_filter in valid_statuses:
        applications = applications.filter(
            status=status_filter
        )
    else:
        status_filter = ""

    application_counts = {
        "total": MembershipApplication.objects.count(),
        "pending": MembershipApplication.objects.filter(
            status=MembershipApplication.Status.PENDING
        ).count(),
        "under_review": MembershipApplication.objects.filter(
            status=MembershipApplication.Status.UNDER_REVIEW
        ).count(),
        "approved": MembershipApplication.objects.filter(
            status=MembershipApplication.Status.APPROVED
        ).count(),
        "rejected": MembershipApplication.objects.filter(
            status=MembershipApplication.Status.REJECTED
        ).count(),
    }

    paginator = Paginator(applications, 20)
    page_obj = paginator.get_page(
        request.GET.get("page")
    )

    return render(
        request,
        "membership/application_list.html",
        {
            "applications": page_obj.object_list,
            "page_obj": page_obj,
            "search_query": search_query,
            "status_filter": status_filter,
            "status_choices": (
                MembershipApplication.Status.choices
            ),
            "application_counts": application_counts,
        },
    )

@staff_member_required
@require_POST
@transaction.atomic
def membership_application_action(request, pk):
    application = get_object_or_404(
        MembershipApplication.objects.select_for_update(),
        pk=pk,
    )

    action = request.POST.get("action", "").strip()

    allowed_actions = {
        "under_review",
        "approve",
        "reject",
    }

    if action not in allowed_actions:
        messages.error(
            request,
            "The requested application action is not valid.",
        )
        return redirect("membership:application_list")

    if (
        application.approved_member_id
        and action != "approve"
    ):
        messages.error(
            request,
            (
                f"Application {application.application_id} has already "
                "created an official member and cannot be rejected or "
                "returned to review."
            ),
        )
        return redirect("membership:application_list")

    if action == "under_review":
        application.status = (
            MembershipApplication.Status.UNDER_REVIEW
        )
        application.reviewed_at = timezone.now()
        application.reviewed_by = request.user

        application.save(
            update_fields=[
                "status",
                "reviewed_at",
                "reviewed_by",
                "updated_at",
            ]
        )

        messages.success(
            request,
            (
                f"Application {application.application_id} was marked "
                "Under Review."
            ),
        )

    elif action == "reject":
        application.status = MembershipApplication.Status.REJECTED
        application.reviewed_at = timezone.now()
        application.reviewed_by = request.user

        application.save(
            update_fields=[
                "status",
                "reviewed_at",
                "reviewed_by",
                "updated_at",
            ]
        )

        messages.success(
            request,
            f"Application {application.application_id} was rejected.",
        )

    elif action == "approve":
        if application.approved_member_id:
            application.status = (
                MembershipApplication.Status.APPROVED
            )
            application.reviewed_at = timezone.now()
            application.reviewed_by = request.user

            application.save(
                update_fields=[
                    "status",
                    "reviewed_at",
                    "reviewed_by",
                    "updated_at",
                ]
            )

            messages.info(
                request,
                (
                    f"Application {application.application_id} was "
                    f"already approved as "
                    f"{application.approved_member.member_id}."
                ),
            )

        else:
            notes = (
                f"Created from membership application "
                f"{application.application_id}.\n"
                f"Applicant email: {application.email}"
            )

            if application.household_information:
                notes += (
                    "\nHousehold information: "
                    f"{application.household_information}"
                )

            if application.message:
                notes += (
                    "\nApplicant message: "
                    f"{application.message}"
                )

            member = Member.objects.create(
                full_name=application.full_name,
                baptism_name=application.baptism_name,
                yenesha_abat=application.yenesha_abat,
                address=application.address,
                phone_number=application.phone_number,
                membership_date=timezone.localdate(),
                is_active=True,
                notes=notes,
            )

            application.status = (
                MembershipApplication.Status.APPROVED
            )
            application.reviewed_at = timezone.now()
            application.reviewed_by = request.user
            application.approved_member = member

            application.save(
                update_fields=[
                    "status",
                    "reviewed_at",
                    "reviewed_by",
                    "approved_member",
                    "updated_at",
                ]
            )

            messages.success(
                request,
                (
                    f"Application {application.application_id} was "
                    f"approved. Member {member.member_id} was created."
                ),
            )

    return redirect("membership:application_list")

@staff_member_required
def member_create(request):
    if request.method == "POST":
        form = MemberForm(request.POST)

        if form.is_valid():
            member = form.save()

            messages.success(
                request,
                f"Member {member.full_name} was added as "
                f"{member.member_id}.",
            )

            return redirect("membership:member_list")
    else:
        form = MemberForm()

    return render(
        request,
        "membership/member_form.html",
        {
            "form": form,
            "page_title": "Add Member",
        },
    )


@staff_member_required
def member_update(request, pk):
    member = get_object_or_404(Member, pk=pk)

    if request.method == "POST":
        form = MemberForm(request.POST, instance=member)

        if form.is_valid():
            form.save()

            messages.success(
                request,
                f"Member {member.member_id} was updated.",
            )

            return redirect("membership:member_list")
    else:
        form = MemberForm(instance=member)

    return render(
        request,
        "membership/member_form.html",
        {
            "form": form,
            "member": member,
            "page_title": "Edit Member",
        },
    )


@staff_member_required
@require_POST
def member_deactivate(request, pk):
    member = get_object_or_404(Member, pk=pk)

    member.is_active = False
    member.save(update_fields=["is_active", "updated_at"])

    messages.success(
        request,
        f"Member {member.member_id} was deactivated.",
    )

    return redirect("membership:member_list")


@staff_member_required
def member_import(request):
    if request.method == "POST":
        form = MemberImportForm(
            request.POST,
            request.FILES,
        )

        if form.is_valid():
            excel_file = form.cleaned_data["excel_file"]

            try:
                workbook = load_workbook(
                    excel_file,
                    data_only=True,
                )
            except Exception:
                messages.error(
                    request,
                    "The uploaded file could not be read. "
                    "Please upload a valid Excel .xlsx file.",
                )
                return redirect("membership:member_import")

            worksheet = workbook.active

            header_values = [
                cell.value
                for cell in worksheet[1]
            ]

            column_indexes = {}

            for index, header in enumerate(header_values):
                normalized = _normalize_header(header)
                field_name = HEADER_ALIASES.get(normalized)

                if field_name and field_name not in column_indexes:
                    column_indexes[field_name] = index

            if "full_name" not in column_indexes:
                messages.error(
                    request,
                    'The spreadsheet must contain a "Full Name" column.',
                )
                return redirect("membership:member_import")

            created_count = 0
            updated_count = 0
            skipped_count = 0

            with transaction.atomic():
                for row_number, row in enumerate(
                    worksheet.iter_rows(
                        min_row=2,
                        values_only=True,
                    ),
                    start=2,
                ):
                    row_data = {}

                    for field_name, column_index in column_indexes.items():
                        if column_index < len(row):
                            row_data[field_name] = row[column_index]
                        else:
                            row_data[field_name] = None

                    if not any(
                        value not in (None, "")
                        for value in row_data.values()
                    ):
                        continue

                    full_name = _clean_text(
                        row_data.get("full_name")
                    )

                    if not full_name:
                        skipped_count += 1
                        continue

                    phone_number = _clean_text(
                        row_data.get("phone_number")
                    )

                    membership_date = _parse_membership_date(
                        row_data.get("membership_date"),
                        workbook.epoch,
                    )

                    defaults = {
                        "full_name": full_name,
                        "baptism_name": _clean_text(
                            row_data.get("baptism_name")
                        ),
                        "address": _clean_text(
                            row_data.get("address")
                        ),
                        "yenesha_abat": _clean_text(
                            row_data.get("yenesha_abat")
                        ),
                        "membership_date": membership_date,
                        "phone_number": phone_number,
                        "is_active": _parse_active_status(
                            row_data.get("is_active")
                        ),
                        "notes": _clean_text(
                            row_data.get("notes")
                        ),
                    }

                    existing_member = None

                    uploaded_member_id = _clean_text(
                        row_data.get("member_id")
                    ).upper()

                    if uploaded_member_id:
                        existing_member = Member.objects.filter(
                            member_id__iexact=uploaded_member_id
                        ).first()

                    if (
                        existing_member is None
                        and phone_number
                    ):
                        existing_member = Member.objects.filter(
                            full_name__iexact=full_name,
                            phone_number=phone_number,
                        ).first()

                    if existing_member:
                        for field_name, value in defaults.items():
                            setattr(existing_member, field_name, value)

                        existing_member.save()
                        updated_count += 1
                    else:
                        # A new KU Member ID is generated automatically.
                        Member.objects.create(**defaults)
                        created_count += 1

            messages.success(
                request,
                (
                    f"Excel import completed: "
                    f"{created_count} created, "
                    f"{updated_count} updated, "
                    f"{skipped_count} skipped."
                ),
            )

            return redirect("membership:member_list")
    else:
        form = MemberImportForm()

    return render(
        request,
        "membership/member_import.html",
        {
            "form": form,
        },
    )

def membership_apply(request):
    if request.method == "POST":
        form = MembershipApplicationForm(request.POST)

        if form.is_valid():
            application = form.save()

            return redirect(
                "membership:application_success",
                application_id=application.application_id,
            )
    else:
        form = MembershipApplicationForm()

    return render(
        request,
        "membership/application_form.html",
        {
            "form": form,
        },
    )


def application_success(request, application_id):
    application = get_object_or_404(
        MembershipApplication,
        application_id=application_id,
    )

    return render(
        request,
        "membership/application_success.html",
        {
            "application": application,
        },
    )
@staff_member_required
def member_export(request):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Members"

    headers = [
        "Member ID",
        "Full Name",
        "Baptism Name",
        "Address",
        "Yenesha Abat",
        "Date",
        "Phone Number",
        "Active Member",
        "Notes",
    ]

    worksheet.append(headers)

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="8F1D2C",
    )

    for cell in worksheet[1]:
        cell.font = Font(
            bold=True,
            color="FFFFFF",
        )
        cell.fill = header_fill
        cell.alignment = Alignment(
            horizontal="center",
        )

    for member in Member.objects.all():
        worksheet.append(
            [
                member.member_id,
                member.full_name,
                member.baptism_name,
                member.address,
                member.yenesha_abat,
                member.membership_date,
                member.phone_number,
                "Yes" if member.is_active else "No",
                member.notes,
            ]
        )

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    column_widths = {
        "A": 16,
        "B": 26,
        "C": 22,
        "D": 38,
        "E": 22,
        "F": 15,
        "G": 20,
        "H": 16,
        "I": 35,
    }

    for column_letter, width in column_widths.items():
        worksheet.column_dimensions[column_letter].width = width

    filename = (
        "st_urael_members_"
        f"{timezone.localdate().isoformat()}.xlsx"
    )

    response = HttpResponse(
        content_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        )
    )

    response["Content-Disposition"] = (
        f'attachment; filename="{filename}"'
    )

    workbook.save(response)

    return response